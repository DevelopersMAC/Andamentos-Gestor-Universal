from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
import time
import uiautomation as uia
import win32gui
import win32con
import win32api
from datetime import date
from selenium import webdriver
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from tkinter import *

ie_options = webdriver.IeOptions()
ie_options.attach_to_edge_chrome = True
ie_options.edge_executable_path = "C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe"


linha = 0

try:
    workbook = load_workbook("base.xlsx")
except:
    messagebox.showinfo(title="Accionar Mac Gestor", message="Base corrompida!")
    exit()

worksheet = workbook.active

def login():

    Login = "san.01125"

    arq = open(r"\\192.168.0.9\Publica\### PUBLICA - MAC BARBOSA ###\00.SITES\01.SANTANDER_FINANCEIRA\GESTOR.txt","r")

    linhas = arq.readlines()

    for linha in linhas:
        Senha = linha

    browser.get("https://www.gestaojudicial.com.br/Paginas/Principal/_FSet_Abertura.asp")

    elementLogin = browser.find_element(By.ID,"txtcd_Logon")

    elementSenha = browser.find_element(By.ID,"txtcd_Pwd")

    elementSeg = browser.find_element(By.ID,"CodSegInformado")

    button = browser.find_element(By.ID,"btOK")

    elementSegCriado = str(browser.find_element(By.ID,"CodSegCriado").get_attribute("value"))

    browser.execute_script('arguments[0].setAttribute("value", arguments[1])', elementLogin, Login)

    browser.execute_script('arguments[0].setAttribute("value", arguments[1])', elementSenha, Senha)

    browser.execute_script('arguments[0].setAttribute("value", arguments[1])', elementSeg, elementSegCriado)

    browser.execute_script('arguments[0].click()', button)

    time.sleep(4)

    browser.switch_to.frame('FraMenu')

    buttonPJ = browser.find_element(By.ID,"PJTxt")

    browser.execute_script('arguments[0].click()', buttonPJ)

def pesquisaCausa():

    time.sleep(4)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraMenu')

    pesqCausa = browser.find_element(By.ID,"PCTxt")

    browser.execute_script('arguments[0].click()', pesqCausa)

    browser.switch_to.default_content()

def inserirContrato(): 

    global linha

    time.sleep(4)

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('fraConsulta')

    browser.switch_to.frame('fraPesquisa')

    inputPesq = browser.find_element(By.ID,"txtPesquisa")

    contrato = worksheet.cell(row=linha + 1, column=1).value

    browser.execute_script("arguments[0].setAttribute('value', arguments[1])", inputPesq, contrato)

    button = browser.find_element(By.ID,"btnPesquisar")

    browser.execute_script("arguments[0].click()", button)

def selecionaCausa():

    time.sleep(4)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('fraConsulta')

    browser.switch_to.frame('fraResultado')

    try:
        tabelas = browser.find_elements(By.TAG_NAME,"tbody")

        linhas = tabelas[1].find_elements(By.TAG_NAME,"tr")

    except:
        
        worksheet.cell(row= linha + 1,column=5).value = "Sem informação"
        workbook.save("base.xlsx")
        browser.quit()
        iniciar()

    time.sleep(2)

    linhaWhite = "Não encontrada"

    for linhaFor in linhas:

        if linhaFor.value_of_css_property('background-Color') == "white":
            
            colunas = linhaFor.find_elements(By.TAG_NAME,"td")

            checkbox = colunas[5].find_element(By.TAG_NAME,"input")

            browser.execute_script("arguments[0].click()", checkbox)

            browser.execute_script("arguments[0].checked = true", checkbox)

            browser.switch_to.default_content()

            browser.switch_to.frame('FraDetalhe')

            browser.switch_to.frame('FraVazio')

            browser.switch_to.frame('fraConsulta')

            browser.switch_to.frame('fraPesquisa')

            img = browser.find_elements(By.TAG_NAME,"img")

            browser.execute_script("arguments[0].click()", img[4])

            time.sleep(2)

            browser.switch_to.default_content()
            
            linhaWhite = "Encontrada"
        
            break
    
    if linhaWhite == "Não encontrada":
        
        worksheet.cell(row= linha + 1,column=5).value = "Sem causa aberta"

        workbook.save("backup.xlsx")

        workbook.save("base.xlsx")

        browser.quit()

        iniciar()

def verificaCadeado():
    
    time.sleep(4)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('Frame_Processo')
    
    divs = browser.find_elements(By.TAG_NAME,"div")

    time.sleep(1)

    img =  divs[2].find_element(By.TAG_NAME,"img")

    time.sleep(1)

    if "leaf_lock.gif" in img.get_attribute('src'):
        
        if "Processo [0]" == divs[7].text:
    
            worksheet.cell(row= linha + 1,column=5).value = "Cadeado Cobrança Fechado e sem cadeado Processo"
        else:
            img =  divs[7].find_element(By.TAG_NAME,"img")

            if "Processo [1]" == divs[7].text and "leaf_open.gif" in img.get_attribute('src'):
            
                links = divs[7].find_element(By.TAG_NAME,"a")

                browser.execute_script("arguments[0].click()", links)

                entrarAndamento()

            elif "Processo [2]" == divs[7].text:

                worksheet.cell(row= linha + 1,column=5).value = "Cadeado processo duplo"

            elif "leaf_lock.gif" in img.get_attribute('src'):

                worksheet.cell(row= linha + 1,column=5).value = "Cadeado cobrança e processo fechado"
            else:
                worksheet.cell(row= linha + 1,column=5).value = "Error"

    elif "leaf_open.gif" in img.get_attribute('src'):
        
        if "Cobrança [0]" == divs[2].text:
    
            worksheet.cell(row= linha + 1,column=5).value = "Sem cadeado cobrança"

        elif "Cobrança [1]" == divs[2].text:
        
            links = divs[2].find_element(By.TAG_NAME,"a")

            browser.execute_script("arguments[0].click()", links)

            entrarAndamento()

        elif "Cobrança [2]" == divs[2].text:

            worksheet.cell(row= linha + 1,column=5).value = "Cadeado processo duplo"

        else:

            worksheet.cell(row= linha + 1,column=5).value = "Error"

def entrarAndamento():

    time.sleep(4)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('fraFichas')
    
    linkAndamento = browser.find_element(By.ID,"AndTxt")

    browser.execute_script("arguments[0].click()", linkAndamento)

    criarAndamento()

def criarAndamento():

    time.sleep(4)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('fraVazio')

    forms = browser.find_elements(By.TAG_NAME,'form')

    nm_Tabela = forms[1].find_element(By.NAME,"nm_Tabela")

    id_Tipo_Andamento = forms[1].find_element(By.NAME,"id_Tipo_Andamento")

    ds_Tipo_Andamento = forms[1].find_element(By.NAME,"ds_Tipo_Andamento")

    browser.execute_script("arguments[0].setAttribute('value', arguments[1])", ds_Tipo_Andamento, "ANDAMENTO")

    browser.execute_script("arguments[0].setAttribute('value', arguments[1])", nm_Tabela, "ANDAMENTO")

    browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Tipo_Andamento, "1")

    button = forms[1].find_element(By.ID,"Button1")

    browser.execute_script("arguments[0].click()", button)

    preencheAndamento()

def verificaPopUP():
    loop = True

    while loop == True:
        msgFWebpage = win32gui.FindWindow(None ,"Mensagem da página da web")

        selecioneUsuario = win32gui.FindWindowEx(msgFWebpage, 0, "Static", "Selecione um usuário responsável.")

        dadosAtualizados = win32gui.FindWindowEx(msgFWebpage, 0, "Static", """Andamento inserido com sucesso.
""")

        btnVBScript = win32gui.FindWindowEx(msgFWebpage, 0, "Button", 'OK')

        if selecioneUsuario > 0:

            win32api.SendMessage(btnVBScript, win32con.BM_CLICK, 0, 0)

            win32api.SendMessage(btnVBScript, win32con.BM_CLICK, 0, 0)

            worksheet.cell(row= linha + 1,column=5).value = "Sem responsavel"

            loop = False

        elif dadosAtualizados > 0:

            win32api.SendMessage(btnVBScript, win32con.BM_CLICK, 0, 0)

            win32api.SendMessage(btnVBScript, win32con.BM_CLICK, 0, 0)

            worksheet.cell(row= linha + 1,column=5).value = "Andamento Inserido com sucesso"

            loop = False

def preencheAndamento():

    time.sleep(4)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('fraVazio')

    browser.switch_to.frame('fraForm')

    form = browser.find_element(By.TAG_NAME,"form")

    responsavel = form.find_element(By.NAME,"cd_Recurso")

    fl_Tipo_Informacao = form.find_element(By.NAME,"fl_Tipo_Informacao")

    if "MAC BARBOSA SOCIEDADE DE ADVOGADOS" in responsavel.text:

        time.sleep(2)

        browser.execute_script("arguments[0].setAttribute('value', arguments[1])", responsavel, "2402")

        browser.execute_script("arguments[0].setAttribute('value', arguments[1])", fl_Tipo_Informacao, "P")

        infoPlan = worksheet.cell(row= linha +1,column=2).value

        sFase = form.find_element(By.NAME,"sFase")

        id_Fase_Andamento = form.find_element(By.NAME,"id_Fase_Andamento")

        if infoPlan == "LOCALIZAÇÃO DO BEM":
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "LOCALIZAÇÃO DO BEM")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "34")

        elif infoPlan == "MANDADO EXPEDIDO - 1º MANDADO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "MANDADO EXPEDIDO - 1º MANDADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "532")

        elif infoPlan == "MANDADO EXPEDIDO - 2º MANDADO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "MANDADO EXPEDIDO - 2º MANDADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "533")    
        
        elif infoPlan == "MANDADO EXPEDIDO - 3º MANDADO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "MANDADO EXPEDIDO - 3º MANDADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "534")

        elif infoPlan == "MANDADO EXPEDIDO - DEMAIS MANDADOS":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "MANDADO EXPEDIDO - DEMAIS MANDADOS")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "535")    

        elif infoPlan == "NÃO LOCALIZAÇÃO DO BEM":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "NÃO LOCALIZAÇÃO DO BEM")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "265")

        elif infoPlan == "FINANCIADO NÃO LOCALIZADO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "FINANCIADO NÃO LOCALIZADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "266")

        elif infoPlan == "FINANCIADO LOCALIZADO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "FINANCIADO LOCALIZADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "267")

        elif infoPlan == "NOTIFICAÇÃO - 1ª NEGATIVA":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "NOTIFICAÇÃO - 1ª NEGATIVA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "395")  

        elif infoPlan == "AGUARDANDO 1ª NOTIFICAÇÃO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "AGUARDANDO 1ª NOTIFICAÇÃO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "408")   

        elif infoPlan == "ANDAMENTO DE COBRANÇA":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "ANDAMENTO DE COBRANÇA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "261")  
        
        elif infoPlan == "VALOR INFERIOR ($)":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "VALOR INFERIOR ($)")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "425")   

        elif infoPlan == "AGUARDANDO CONTRATO ORIGINAL":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "AGUARDANDO CONTRATO ORIGINAL")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "410") 

        elif infoPlan == "Aguardando Termo de Subst de Garantia":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "Aguardando Termo de Subst de Garantia")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "451")   

        elif infoPlan == "CALCULADORA NEGATIVA":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "CALCULADORA NEGATIVA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "414")  

        elif infoPlan == "FALTA NOTA FISCAL":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "FALTA NOTA FISCAL")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "417")    

        elif infoPlan == "FORA DA PARAMETRIZAÇÃO":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "FORA DA PARAMETRIZAÇÃO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "419")  
        
        elif infoPlan == "SUSPEITA DE FRAUDE":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "SUSPEITA DE FRAUDE")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "349")  

        elif infoPlan == "AGUARDANDO NOTIFICAÇÃO DA PRAÇA":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "AGUARDANDO NOTIFICAÇÃO DA PRAÇA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "411")  

        elif infoPlan == "RESOLVIDO (RETOMADA/RECEBIMENTO)":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "RESOLVIDO (RETOMADA/RECEBIMENTO)")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "422") 

        elif infoPlan == "SEM IMAGEM DO CONTRATO - Terceirizada":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "SEM IMAGEM DO CONTRATO - Terceirizada")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "423") 

        elif infoPlan == "NÃO DEFINIDA":

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "NÃO DEFINIDA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "0")

        elif infoPlan == "Aguardando Deferimento da Liminar":
    
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "Aguardando Deferimento da Liminar")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "452")

        elif infoPlan == "LIMINAR DEFERIDA":
        
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "LIMINAR DEFERIDA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "277")

        elif infoPlan == "LIMINAR INDEFERIDA":
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "LIMINAR INDEFERIDA")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "278")

        elif infoPlan == "AGUARDANDO EXPEDICAO DE MANDADO":
                
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "AGUARDANDO EXPEDICAO DE MANDADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "405")  

        elif infoPlan == "MANDADO DE BUSCA E APREENSÃO / REINTEGRAÇÃO DE POSSE EXPEDIDO":
                
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "MANDADO DE BUSCA E APREENSÃO / REINTEGRAÇÃO DE POSSE EXPEDIDO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "400")

        elif infoPlan == "AGUARDANDO CUMPRIMENTO DO MANDADO":
                
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "AGUARDANDO CUMPRIMENTO DO MANDADO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "454") 

        elif infoPlan == "MANDADO NEGATIVO":
                
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "MANDADO NEGATIVO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "291")

        elif infoPlan == "REQUERIDA EXPEDIÇÃO DE OFÍCIOS":
                    
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "REQUERIDA EXPEDIÇÃO DE OFÍCIOS")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "280")

        elif infoPlan == "PROCESSO EXTINTO S/ JULGAMENTO DO MÉRITO":
                    
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "PROCESSO EXTINTO S/ JULGAMENTO DO MÉRITO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "325")

        elif infoPlan == "CONVERSÃO DA BUSCA E APREENSÃO EM EXECUÇÃO":
                    
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", sFase, "CONVERSÃO DA BUSCA E APREENSÃO EM EXECUÇÃO")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", id_Fase_Andamento, "363")                                                                                    

        else:
            worksheet.cell(row= linha + 1,column=5).value = "Descrição incorreta"
            exit()

        data = worksheet.cell(row=linha + 1,column=4).value

        ds_Andamento = form.find_element(By.NAME,"ds_Andamento")

        dt_Andamento = form.find_element(By.NAME,"dt_Andamento")
        
        if data == None:

            data = date.today().strftime("%d/%m/%Y")
            
            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", dt_Andamento, data)
        else:

            browser.execute_script("arguments[0].setAttribute('value', arguments[1])", dt_Andamento, data.strftime("%d/%m/%Y"))
        
        browser.execute_script("arguments[0].setAttribute('value', arguments[1])", ds_Andamento, worksheet.cell(row=linha + 1,column=3).value)

        fl_Status = form.find_elements(By.ID,"fl_Status")

        fl_Tipo_Informacao = form.find_elements(By.ID,"fl_Tipo_Informacao")

        browser.execute_script("arguments[0].setAttribute('value', arguments[1])", fl_Status[0],"C")

        browser.execute_script("arguments[0].setAttribute('Selected', arguments[1])", fl_Status[0],True)

        browser.execute_script("arguments[0].setAttribute('Selected', arguments[1])", fl_Tipo_Informacao[0], True)

        incluir = form.find_element(By.NAME,"btIncluir")

        browser.execute_script("arguments[0].click()",incluir)

        time.sleep(4)

        verificaPopUP()
        
    else:
        worksheet.cell(row= linha + 1,column=5).value = "Sem Responsavel"

def iniciar():

    global linha, browser, ie_options

    browser = webdriver.Ie(options=ie_options)
    
    login()

    while worksheet.cell(row= linha + 1,column=1).value != None:
        
        if worksheet.cell(row= linha + 1,column=5).value == None:

            pesquisaCausa()

            inserirContrato()

            selecionaCausa()

            verificaCadeado()

            linha = linha + 1

            workbook.save("base.xlsx")
            
            workbook.save("backup.xlsx")

        else:
            linha = linha + 1

try:
    
    iniciar()
    messagebox.showinfo(title="Mac Gestor", message="Fim")
    
except:
    workbook.save("base.xlsx")
    browser.quit()
    iniciar()